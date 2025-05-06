from flask import Blueprint, render_template, request, flash, url_for, send_file, current_app, session, redirect 
import pandas as pd
import os
import math
from datetime import datetime, timedelta, time
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

programacion_bp = Blueprint('programacion', __name__, template_folder='templates')

# Carpeta y constantes
UPLOAD_FOLDER    = os.path.abspath(os.path.dirname(__file__))
CONTRACT_HOURS   = {'24HS':6,'30HS':6,'35HS':7,'36HS':6}
SERVICE_KEY_MAP  = {
    'Sop_Conectividad':'Internet','Sop_Flow':'Flow','Esp_CATV':'CATV',
    'Esp_Movil':'Movil','Esp_XDSL':'XDSL','Digital':'Digital',
    'CBS':'CBS','SMB_TecnicaIN':'TecnicaIN','SMB_Digital':'Digital'
}

@programacion_bp.route('/programacion', methods=['GET','POST'])
def programacion():
    nomina_path = session.get('nomina_path')
    if not nomina_path or not os.path.exists(nomina_path):
        flash('Antes de generar programación, sube la nómina.', 'warning')
        return redirect(url_for('upload_nomina'))

    download_url = None
    if request.method=='POST':
        servicio   = request.form.get('servicio')
        req_file   = request.files.get('requeridos')
        if not servicio or not req_file:
            flash('Selecciona servicio y archivo de requeridos.', 'warning')
            return render_template('programacion.html')

        req_path   = os.path.join(UPLOAD_FOLDER,'requeridos_temp.xlsx')
        req_file.save(req_path)
        df_nom     = pd.read_excel(nomina_path)
        df_nom.columns = df_nom.columns.str.strip()

        df_d = pd.read_excel(req_path, sheet_name=servicio, skiprows=[0,2], header=0)
        df_d.rename(columns={df_d.columns[0]:'Intervalo'},inplace=True)
        df_d['Intervalo']=pd.to_datetime(df_d['Intervalo'],format='%H:%M:%S',errors='coerce').dt.time
        df_d.dropna(subset=['Intervalo'],inplace=True)
        date_cols=[c for c in df_d.columns[1:] if not pd.isna(pd.to_datetime(str(c),errors='coerce'))]
        df_long = (df_d.melt(
            id_vars=['Intervalo'],value_vars=date_cols,
            var_name='Fecha',value_name='Requeridos'
        ).dropna(subset=['Requeridos']))
        df_long['Fecha']=pd.to_datetime(df_long['Fecha']).dt.date
        df_long['Requeridos']=df_long['Requeridos'].astype(int)

        key = SERVICE_KEY_MAP.get(servicio,servicio)
        df_x = df_nom[
            df_nom['SERVICIO'].str.contains(key,case=False,na=False)&
            (df_nom['ACTIVO'].str.upper()=='ACTIVO')
        ].copy()
        df_x['INGRESO'] = pd.to_datetime(df_x['INGRESO'],format='%H:%M:%S',errors='coerce').dt.time
        df_x['EGRESO']  = [
            (datetime.combine(datetime.today(),ing)+timedelta(hours=CONTRACT_HOURS.get(str(con).strip().upper(),24))).time()
            for ing,con in zip(df_x['INGRESO'],df_x['CONTRATO'])
        ]

        def assign_off_days(df):
            off_counts={'24HS':3,'30HS':2,'35HS':2,'36HS':1}
            offs=[]
            for idx,row in df.iterrows():
                c=str(row['CONTRATO']).strip().upper()
                d=off_counts.get(c,0)
                offs.append([(idx+k)%7 for k in range(d)])
            return offs
        df_x['OFF_DAYS']=assign_off_days(df_x)

        rows=[]
        for _,r in df_long.iterrows():
            f,i,req=r['Fecha'],r['Intervalo'],r['Requeridos']
            if req<10:    li,up=max(req-1,0),req+1
            elif req<20:  li,up=max(req-2,0),req+2
            else:         li,up=math.floor(req*0.9),math.ceil(req*1.1)
            prime='Prime' if time(9,0)<=i<time(21,0) else 'No prime'
            active=df_x[~df_x['OFF_DAYS'].apply(lambda offs: f.weekday() in offs)]
            norm=(active['EGRESO']>=active['INGRESO'])&(active['INGRESO']<=i)&(i<=active['EGRESO'])
            wrap=(active['EGRESO']<active['INGRESO'])&(active['INGRESO']<=i)
            pres=active[norm|wrap]
            cnt=len(pres)
            leader_col=next((c for c in pres.columns if c.strip().lower() in ['jefe','nuevo superior','lider']),None)
            lideres=pres[leader_col].dropna().unique().tolist() if leader_col else []
            faltante=max(li-cnt,0)
            rows.append({
                'Fecha':f,'Intervalo':i.strftime('%H:%M'),'Prime':prime,
                'Requeridos':req,'Limite Inferior':li,'Limite Superior':up,
                'Faltante':faltante,'Asignados':cnt,'Estado':'',
                'Lider':';'.join(lideres),
                'DNI_Presentes':';'.join(pres['DNI'].astype(str)),
                'Nombres_Presentes':';'.join(pres['NOMBRE'])
            })

        df_out=pd.DataFrame(rows)
        order=['Fecha','Intervalo','Prime','Requeridos',
               'Limite Inferior','Limite Superior','Faltante',
               'Asignados','Estado','Lider',
               'DNI_Presentes','Nombres_Presentes']
        df_out=df_out[order]
        df_out['Fecha']=pd.to_datetime(df_out['Fecha'])

        out_path=os.path.join(UPLOAD_FOLDER,'reporte.xlsx')
        red=PatternFill('solid', fgColor='FF0000')
        green=PatternFill('solid', fgColor='00FF00')
        yellow=PatternFill('solid', fgColor='FFFF00')
        orange=PatternFill('solid', fgColor='FFA500')

        with pd.ExcelWriter(out_path,engine='openpyxl') as w:
            df_out.to_excel(w,'Nomina',index=False)
            ws=w.sheets['Nomina']
            for idx,col in enumerate(order, start=1):
                ws.cell(1,idx).font=Font(bold=True)
                ws.column_dimensions[get_column_letter(idx)].width=13
            for c in ws['A'][1:]: c.number_format='YYYY-MM-DD'

            estado_col = order.index('Estado') + 1
            asignados_col = order.index('Asignados') + 1
            li_col = order.index('Limite Inferior') + 1
            up_col = order.index('Limite Superior') + 1

            for r in range(2, ws.max_row + 1):
                asignados = ws.cell(r, asignados_col).value
                limite_inf = ws.cell(r, li_col).value
                limite_sup = ws.cell(r, up_col).value

                cell = ws.cell(r, estado_col)
                if asignados < limite_inf:
                    cell.fill = red
                    cell.value = 'UNDER'
                elif asignados > limite_sup:
                    cell.fill = yellow
                    cell.value = 'OVER'
                elif asignados == limite_inf:
                    cell.fill = orange
                else:
                    cell.fill = green

            falt=df_out[df_out['Faltante']>0]
            if not falt.empty:
                msgs=[f"{row['Fecha'].date()} {row['Intervalo']}" for _,row in falt.iterrows()]
                msg="Faltantes en: "+", ".join(msgs)
                nc=len(order)+1
                ws.cell(1,nc,msg).font=Font(bold=True,color='FF0000')
                ws.column_dimensions[get_column_letter(nc)].width=len(msg)/1.2

        download_url=url_for('programacion.download',servicio=servicio)

    return render_template('programacion.html', download_url=download_url)

@programacion_bp.route('/programacion/download')
def download():
    servicio=request.args.get('servicio','programacion')
    key=SERVICE_KEY_MAP.get(servicio,servicio)
    filename=f"programacion - {key}.xlsx"
    return send_file(os.path.join(UPLOAD_FOLDER,'reporte.xlsx'),
                     as_attachment=True, download_name=filename)
